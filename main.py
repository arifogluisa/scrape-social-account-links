from openpyxl import Workbook                 # we write the result to Excel file with using this package,
from openpyxl.utils import get_column_letter  # use this method to get header of Excel file
from openpyxl.styles import Font              # and make style of header bold

#######################################################################################################################
# import each method for each website
from scrapper.binance import binance_scraper
from scrapper.adsoftheworld import adsoft_scraper
from scrapper.tether import tether_scraper
from scrapper.ethereum import ethereum_scrapper
from scrapper.avax import avax_scrapper
from scrapper.dogecoin import dogecoin_scrapper
from scrapper.terramoney import terra_money
from scrapper.mashable import mashable
from scrapper.litecoin import litecoin
from scrapper.appypie import appypie
from scrapper.catcoin import catcoin
from scrapper.doxedtoken import doxedtoken
from scrapper.markmeta import markmeta
#######################################################################################################################

wb = Workbook()
ws = wb.active
ws.title = 'Task'
headings = ['URLs', 'Reddit', 'Telegram', 'Discord', 'Medium', 'Pinterest']
ws.append(headings)

# we store all scraping result of websites in a list called 'all_lists'
all_lists = [tether_scraper(), adsoft_scraper(), binance_scraper(), ethereum_scrapper(),
             avax_scrapper(), dogecoin_scrapper(), terra_money(), mashable(), litecoin(),
             appypie(), catcoin(), doxedtoken(), markmeta()]

# and write a for loop to write data to Excel file
for each_list in all_lists:
    ws.append(each_list)
# make some style for header of Excel file
for col in range(1, 8):
    ws[get_column_letter(col) + '1'].font = Font(bold=True)
# finally, create and save Excel file
wb.save('task.xlsx')
